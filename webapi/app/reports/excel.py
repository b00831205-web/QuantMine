"""Excel report export: write raw tables from the report context into .xlsx."""

from __future__ import annotations

import io

from openpyxl import Workbook


def _append_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: list[list[object]],
) -> None:
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    # 简单列宽：避免内容挤在一起
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 14


def build_xlsx(context: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # 删掉默认空 sheet，统一用下面创建的

    _append_sheet(
        wb,
        "Summary",
        ["Field", "Value"],
        [[key, value] for key, value in context["meta"].items()],
    )

    ic_rows = context["ic_rows"]
    if ic_rows:
        headers = list(ic_rows[0].keys())
        _append_sheet(
            wb,
            "IC",
            headers,
            [[row.get(header, "") for header in headers] for row in ic_rows],
        )

    groups = context["backtest_groups"]
    if groups:
        backtest_headers = [
            "group", "annual_return", "volatility", "sharpe",
            "max_drawdown", "win_rate", "turnover",
        ]
        backtest_rows: list[list[object]] = []
        for group in groups:
            backtest_rows.append([group["label"], "", "", "", "", "", ""])
            for row in group["rows"]:
                backtest_rows.append(
                    [row["group"], row["ann"], row["vol"], row["sharpe"],
                     row["mdd"], row["win"], row["turnover"]]
                )
        _append_sheet(wb, "Backtest", backtest_headers, backtest_rows)

        mono_headers = ["group", "mono_corr", "mono_p", "mono_daily", "mono_pos"]
        mono_rows = [
            [group["label"], group["mono"]["corr"], group["mono"]["p"],
             group["mono"]["daily"], group["mono"]["pos"]]
            for group in groups
        ]
        _append_sheet(wb, "Monotonicity", mono_headers, mono_rows)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()